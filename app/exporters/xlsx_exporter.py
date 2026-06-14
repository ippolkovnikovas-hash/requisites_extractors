"""
Экспорт реквизитов в Excel.
Лист 1 — данные по строкам (русское название → значение → статус валидации).
Лист 2 — сырой отчёт валидации.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.schemas.requisites import RequisitesData
from app.schemas.validation import ValidationReport


# Маппинг python-имя → человекочитаемое название (порядок совпадает с shablon.docx)
_FIELD_LABELS: list[tuple[str, str]] = [
    ("company_name",          "Полное наименование организации"),
    ("short_name",            "Сокращённое наименование"),
    ("legal_address",         "Юридический адрес"),
    ("postal_address",        "Почтовый адрес"),
    ("ogrn",                  "ОГРН / ОГРНИП"),
    ("inn",                   "ИНН"),
    ("kpp",                   "КПП"),
    ("bank_name",             "Наименование банка"),
    ("checking_account",      "Расчётный счёт (Р/счет)"),
    ("correspondent_account", "Корреспондентский счёт (К/счет)"),
    ("bik",                   "БИК"),
    ("ceo_position",          "Должность руководителя"),
    ("ceo_fio_full",          "ФИО руководителя (полностью)"),
    ("ceo_fio",               "ФИО руководителя (кратко)"),
    ("phone",                 "Телефон"),
    ("email",                 "Электронная почта"),
]

# Поля с алгоритмической валидацией
_VALIDATED_FIELDS = {"inn", "kpp", "ogrn", "bik", "checking_account", "correspondent_account"}

_GREEN  = PatternFill("solid", fgColor="D6F5D6")
_RED    = PatternFill("solid", fgColor="FAD7D7")
_YELLOW = PatternFill("solid", fgColor="FFF3CD")
_HEADER = PatternFill("solid", fgColor="2E5090")
_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _status(field_name: str, value: str | None, validation: ValidationReport) -> tuple[str, PatternFill]:
    if value is None:
        return "не найдено", _YELLOW
    if field_name in _VALIDATED_FIELDS:
        fv = getattr(validation, field_name, None)
        if fv is None:
            return "ок (не проверялось)", _GREEN
        return ("ок", _GREEN) if fv.valid else (f"ошибка: {fv.reason}", _RED)
    return "ок", _GREEN


def export_xlsx(
    document_id: str,
    requisites: RequisitesData,
    validation: ValidationReport,
) -> Path:
    out_path = settings.exports_folder / f"{document_id}_result.xlsx"
    wb = Workbook()

    # ── Лист 1: реквизиты ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Реквизиты"

    headers = ["Поле", "Значение", "Статус валидации"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN

    data_dict = requisites.model_dump()

    for row_idx, (field_name, label) in enumerate(_FIELD_LABELS, start=2):
        value = data_dict.get(field_name)
        status_text, fill = _status(field_name, value, validation)

        c_label = ws.cell(row=row_idx, column=1, value=label)
        c_label.font = Font(bold=True)
        c_label.border = _THIN
        c_label.alignment = Alignment(wrap_text=True, vertical="center")

        c_value = ws.cell(row=row_idx, column=2, value=value or "")
        c_value.border = _THIN
        c_value.alignment = Alignment(wrap_text=True, vertical="center")

        c_status = ws.cell(row=row_idx, column=3, value=status_text)
        c_status.fill = fill
        c_status.border = _THIN
        c_status.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.row_dimensions[1].height = 22

    # ── Лист 2: отчёт валидации ─────────────────────────────────────────
    ws2 = wb.create_sheet("Валидация")
    ws2.append(["Поле", "Валидно", "Значение", "Причина ошибки"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER
        cell.border = _THIN

    validated = ["inn", "kpp", "ogrn", "bik", "checking_account", "correspondent_account"]
    for fname in validated:
        fv = getattr(validation, fname, None)
        if fv:
            fill = _GREEN if fv.valid else _RED
            row = [fname, "Да" if fv.valid else "Нет", fv.value or "", fv.reason or ""]
            ws2.append(row)
            for c in ws2[ws2.max_row]:
                c.border = _THIN
                c.fill = fill

    if validation.errors:
        ws2.append([])
        ws2.append(["Ошибки / кросс-проверки:"])
        for err in validation.errors:
            ws2.append(["", "", err])

    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 30

    wb.save(out_path)
    return out_path